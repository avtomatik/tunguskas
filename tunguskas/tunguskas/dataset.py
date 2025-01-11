#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Oct 20 21:47:18 2022

@author: alexandermikhailov
"""

from pathlib import Path

import rarfile
from openpyxl import load_workbook

# =============================================================================
# TODO: https://git-lfs.com
# =============================================================================
FILE_NAME = 'urovni_p_i_n_tunguski.rar'


with rarfile.RarFile(
    (
        Path(__file__).parent.parent
        .joinpath('data')
        .joinpath('raw')
        .joinpath(FILE_NAME)
    )
) as rf:
    for f in rf.namelist():
        print(f)

    workbook = load_workbook(
        rf.open('!Уровни постов.xlsx'),
        # =====================================================================
        # TODO: Fix the Bug Below
        # =====================================================================
        # =====================================================================
        #     raise BadRarFile("Failed the read enough data: req=%d got=%d" % (orig, len(data)))
        # rarfile.BadRarFile: Failed the read enough data: req=8331 got=51
        # =====================================================================
        read_only=True,
        keep_links=False
    )
    sheet_names = workbook.sheetnames
    print(sheet_names)
